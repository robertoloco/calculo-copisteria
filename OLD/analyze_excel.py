import openpyxl
import json

wb = openpyxl.load_workbook('calculo tarifas.xlsx')
print('=== HOJAS ===')
print(wb.sheetnames)
print()

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n=== HOJA: {sn} ===')
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows[:20]):
        print(f'Fila {i+1}: {row}')
